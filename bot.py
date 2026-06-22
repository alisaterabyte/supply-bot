"""
Supply Tracking Telegram Bot
"""

import asyncio
import logging
import os
from datetime import datetime
import zoneinfo

NY_TZ = zoneinfo.ZoneInfo("America/New_York")

def now_ny() -> str:
    return datetime.now(NY_TZ).strftime("%Y-%m-%d %H:%M")
from pathlib import Path

import openpyxl
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
EXCEL_FILE = Path(os.environ.get("DATA_DIR", str(Path(__file__).parent))) / "stock.xlsx"

PRODUCTS = [
    {"name": "SEALING TAPE",           "unit": "box"},
    {"name": "FOIL BUBBLE MAILERS",    "unit": "box"},
    {"name": "CARDBOARD SMALL",        "unit": "pallet"},
    {"name": "PALLET WRAP",            "unit": "box"},
    {"name": "8X6X4 FOR TEMU",         "unit": "pallet"},
    {"name": "10X6X6 FOR TEMU",        "unit": "pallet"},
    {"name": "ICE",                    "unit": "pallet"},
    {"name": "BUBBLE MAILERS",         "unit": "box"},
    {"name": "KRAFT PAPER",            "unit": "rolls"},
    {"name": "SMALL PRINTER STICKERS", "unit": "box"},
    {"name": "RIBBON STICKERS",        "unit": "rolls"},
    {"name": "ROLL FILM",              "unit": "kg"},
]

LOCK = asyncio.Lock()


# ── Excel helpers ─────────────────────────────────────────────────────────────

def init_excel():
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        changed = False
        if "Usage" not in wb.sheetnames:
            ws2 = wb.create_sheet("Usage")
            ws2.append(["Date", "Period", "Product", "Amount", "Unit"])
            changed = True
        if "History" not in wb.sheetnames:
            wsh = wb.create_sheet("History")
            wsh.append(["Date", "User", "Product", "Action", "Old value", "New value", "Unit"])
            changed = True
        if changed:
            wb.save(EXCEL_FILE)
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["Product", "Unit", "Quantity"])
    for p in PRODUCTS:
        ws.append([p["name"], p["unit"], 0])
    ws2 = wb.create_sheet("Usage")
    ws2.append(["Date", "Period", "Product", "Amount", "Unit"])
    wsh = wb.create_sheet("History")
    wsh.append(["Date", "User", "Product", "Action", "Old value", "New value", "Unit"])
    wb.save(EXCEL_FILE)
    log.info("Created %s", EXCEL_FILE)


def get_quantities() -> dict:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Stock"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                result[row[0]] = float(row[2] or 0)
            except (TypeError, ValueError):
                result[row[0]] = 0
    return result


def update_quantity(product_name: str, new_qty: float, unit: str,
                    user: str, action: str, old_qty: float):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Stock"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_name:
            row[2].value = new_qty
            wsh = wb["History"]
            wsh.append([
                now_ny(),
                user, product_name, action,
                old_qty, new_qty, unit,
            ])
            wb.save(EXCEL_FILE)
            return
    raise ValueError(f"Product '{product_name}' not found")


def get_history(limit: int = 20) -> list:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["History"]
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
    return rows[-limit:]


def add_usage_entry(period: str, product_name: str, amount: str, unit: str) -> bool:
    """Returns True if new entry, False if existing entry was updated."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Usage"]
    for row in ws.iter_rows(min_row=2):
        if row[1].value == period and row[2].value == product_name:
            row[0].value = now_ny()
            row[3].value = amount
            wb.save(EXCEL_FILE)
            return False
    ws.append([now_ny(), period, product_name, amount, unit])
    wb.save(EXCEL_FILE)
    return True


def get_usage_entries() -> list:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Usage"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(row)
    return rows


# ── FSM States ────────────────────────────────────────────────────────────────

class Stock(StatesGroup):
    action = State()
    product = State()
    quantity = State()


class Usage(StatesGroup):
    action = State()      # view or add
    product = State()
    period = State()
    amount = State()


# ── Keyboards ─────────────────────────────────────────────────────────────────

def kb_main() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📦 View stock",      callback_data="view")],
        [InlineKeyboardButton(text="✏️ Update stock",    callback_data="update")],
        [InlineKeyboardButton(text="🕓 View history",  callback_data="history")],
    ])


def kb_action() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Add (incoming)", callback_data="action_add")],
        [InlineKeyboardButton(text="➖ Remove (used)",  callback_data="action_sub")],
        [InlineKeyboardButton(text="🔄 Set exact value", callback_data="action_set")],
        [InlineKeyboardButton(text="⬅️ Back",           callback_data="back_main")],
    ])


def kb_products(back_cb: str = "back_action") -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(
            text=f"{p['name']} ({p['unit']})",
            callback_data=f"product_{i}",
        )]
        for i, p in enumerate(PRODUCTS)
    ]
    buttons.append([InlineKeyboardButton(text="⬅️ Back", callback_data=back_cb)])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def kb_cancel(cb: str = "back_main") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Cancel", callback_data=cb)],
    ])


def kb_usage_menu() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 View usage log",  callback_data="usage_view")],
        [InlineKeyboardButton(text="✏️ Add usage entry", callback_data="usage_add")],
    ])


def kb_usage_products() -> InlineKeyboardMarkup:
    return kb_products(back_cb="usage_back_menu")


def kb_cancel_usage() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Cancel", callback_data="usage_back_menu")],
    ])


# ── Helpers ───────────────────────────────────────────────────────────────────

def qty_str(v: float) -> str:
    return str(int(v)) if v == int(v) else str(v)


def format_stock(quantities: dict) -> str:
    lines = ["📦 *Current stock:*\n"]
    for p in PRODUCTS:
        qty = quantities.get(p["name"], 0)
        lines.append(f"• {p['name']}: *{qty_str(qty)}* {p['unit']}")
    return "\n".join(lines)


def format_usage(entries: list) -> str:
    if not entries:
        return "📊 *Usage log is empty.*\n\nAdd entries with /usage"
    lines = ["📊 *Usage log:*\n"]
    for row in entries:
        date_added, period, product, amount, unit = row
        lines.append(f"• {period} | {product}: {amount} {unit or ''}")
    return "\n".join(lines)


# ── Handlers ──────────────────────────────────────────────────────────────────

def setup_handlers(dp: Dispatcher) -> None:

    # ── /start ────────────────────────────────────────────────────────────────

    @dp.message(CommandStart())
    async def cmd_start(message: Message, state: FSMContext) -> None:
        await state.clear()
        await message.answer(
            "👋 Welcome to *Supply Tracker*!\n\nWhat would you like to do?",
            parse_mode="Markdown",
            reply_markup=kb_main(),
        )

    # ── Stock: view ───────────────────────────────────────────────────────────

    @dp.callback_query(F.data == "view")
    async def cb_view(call: CallbackQuery) -> None:
        await call.answer()
        async with LOCK:
            quantities = await asyncio.to_thread(get_quantities)
        await call.message.edit_text(
            format_stock(quantities), parse_mode="Markdown", reply_markup=kb_main()
        )

    @dp.callback_query(F.data == "update")
    async def cb_update(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        await state.set_state(Stock.action)
        await call.message.edit_text("Choose action:", reply_markup=kb_action())

    @dp.callback_query(F.data == "back_main")
    async def cb_back_main(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        await state.clear()
        await call.message.edit_text(
            "👋 Welcome to *Supply Tracker*!\n\nWhat would you like to do?",
            parse_mode="Markdown",
            reply_markup=kb_main(),
        )

    @dp.callback_query(F.data == "history")
    async def cb_history(call: CallbackQuery) -> None:
        await call.answer()
        rows = await asyncio.to_thread(get_history, 20)
        if not rows:
            text = "🕓 *Change history*\n\nNo changes yet."
        else:
            lines = ["🕓 *Change history (last 20):*\n"]
            for row in reversed(rows):
                date, user, product, action, old_val, new_val, unit = row
                lines.append(
                    f"• `{date}` {user}\n"
                    f"  {action} *{product}*: {qty_str(float(old_val or 0))} → {qty_str(float(new_val or 0))} {unit}"
                )
            text = "\n".join(lines)
        await call.message.edit_text(text, parse_mode="Markdown", reply_markup=kb_main())

    # ── Stock: action/product/quantity ────────────────────────────────────────

    @dp.callback_query(F.data.in_({"action_add", "action_sub", "action_set"}))
    async def cb_action(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        action = call.data.replace("action_", "")
        await state.update_data(action=action)
        await state.set_state(Stock.product)
        await call.message.edit_text("Select product:", reply_markup=kb_products())

    @dp.callback_query(F.data == "back_action")
    async def cb_back_action(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        await state.set_state(Stock.action)
        await call.message.edit_text("Choose action:", reply_markup=kb_action())

    @dp.callback_query(Stock.product, F.data.startswith("product_"))
    async def cb_product_stock(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        idx = int(call.data.split("_")[1])
        product = PRODUCTS[idx]
        data = await state.get_data()
        action = data.get("action", "add")
        await state.update_data(product_idx=idx)
        await state.set_state(Stock.quantity)

        action_text = {
            "add": "➕ How many to *add*",
            "sub": "➖ How many to *remove*",
            "set": "🔄 Set exact quantity to",
        }[action]

        await call.message.edit_text(
            f"{action_text} for *{product['name']}*?\n"
            f"Unit: *{product['unit']}*\n\nEnter a number:",
            parse_mode="Markdown",
            reply_markup=kb_cancel(),
        )

    @dp.message(Stock.quantity)
    async def on_quantity(message: Message, state: FSMContext) -> None:
        text = message.text.strip().replace(",", ".")
        try:
            amount = float(text)
            if amount < 0:
                raise ValueError
        except ValueError:
            await message.answer("⚠️ Please enter a valid positive number:", reply_markup=kb_cancel())
            return

        data = await state.get_data()
        action = data["action"]
        product = PRODUCTS[data["product_idx"]]

        user = message.from_user
        user_label = f"@{user.username}" if user.username else user.full_name

        async with LOCK:
            quantities = await asyncio.to_thread(get_quantities)
            current = quantities.get(product["name"], 0)
            new_qty = (
                current + amount if action == "add" else
                max(0, current - amount) if action == "sub" else
                amount
            )
            action_label = {"add": "➕ Add", "sub": "➖ Remove", "set": "🔄 Set"}[action]
            await asyncio.to_thread(
                update_quantity, product["name"], new_qty,
                product["unit"], user_label, action_label, current,
            )

        action_emoji = {"add": "➕", "sub": "➖", "set": "🔄"}[action]
        await message.answer(
            f"{action_emoji} *{product['name']}* updated!\n"
            f"{qty_str(current)} → *{qty_str(new_qty)}* {product['unit']}",
            parse_mode="Markdown",
            reply_markup=kb_main(),
        )
        await state.clear()

    # ── /usage ────────────────────────────────────────────────────────────────

    @dp.message(Command("usage"))
    async def cmd_usage(message: Message, state: FSMContext) -> None:
        await state.clear()
        await message.answer(
            "📊 *Usage Tracker*\n\nTrack how much of each product was used over a time period.",
            parse_mode="Markdown",
            reply_markup=kb_usage_menu(),
        )

    @dp.callback_query(F.data == "usage_view")
    async def cb_usage_view(call: CallbackQuery) -> None:
        await call.answer()
        async with LOCK:
            entries = await asyncio.to_thread(get_usage_entries)
        await call.message.edit_text(
            format_usage(entries), parse_mode="Markdown", reply_markup=kb_usage_menu()
        )

    @dp.callback_query(F.data == "usage_add")
    async def cb_usage_add(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        await state.set_state(Usage.product)
        await call.message.edit_text("Select product:", reply_markup=kb_usage_products())

    @dp.callback_query(F.data == "usage_back_menu")
    async def cb_usage_back_menu(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        await state.clear()
        await call.message.edit_text(
            "📊 *Usage Tracker*\n\nTrack how much of each product was used over a time period.",
            parse_mode="Markdown",
            reply_markup=kb_usage_menu(),
        )

    @dp.callback_query(Usage.product, F.data.startswith("product_"))
    async def cb_product_usage(call: CallbackQuery, state: FSMContext) -> None:
        await call.answer()
        idx = int(call.data.split("_")[1])
        await state.update_data(product_idx=idx)
        await state.set_state(Usage.period)
        await call.message.edit_text(
            f"*{PRODUCTS[idx]['name']}* selected.\n\n"
            f"Enter the time period (e.g. *June*, *Week 1*, *Q2 2025*):",
            parse_mode="Markdown",
            reply_markup=kb_cancel_usage(),
        )

    @dp.message(Usage.period)
    async def on_usage_period(message: Message, state: FSMContext) -> None:
        await state.update_data(period=message.text.strip())
        data = await state.get_data()
        product = PRODUCTS[data["product_idx"]]
        await state.set_state(Usage.amount)
        await message.answer(
            f"Period: *{data['period']}*\n"
            f"Product: *{product['name']}* ({product['unit']})\n\n"
            f"Enter amount used (number or note, e.g. *3*, *1.5*, *~2 boxes*):",
            parse_mode="Markdown",
            reply_markup=kb_cancel_usage(),
        )

    @dp.message(Usage.amount)
    async def on_usage_amount(message: Message, state: FSMContext) -> None:
        amount = message.text.strip()
        data = await state.get_data()
        product = PRODUCTS[data["product_idx"]]
        period = data["period"]

        async with LOCK:
            is_new = await asyncio.to_thread(
                add_usage_entry, period, product["name"], amount, product["unit"]
            )

        label = "✅ Usage logged!" if is_new else "🔄 Entry updated!"
        await message.answer(
            f"{label}\n\n"
            f"Period: *{period}*\n"
            f"Product: *{product['name']}*\n"
            f"Amount: *{amount}* {product['unit']}",
            parse_mode="Markdown",
            reply_markup=kb_usage_menu(),
        )
        await state.clear()


# ── Main ──────────────────────────────────────────────────────────────────────

async def main() -> None:
    init_excel()
    bot = Bot(token=TELEGRAM_TOKEN)
    dp = Dispatcher(storage=MemoryStorage())
    setup_handlers(dp)
    log.info("Bot started")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
